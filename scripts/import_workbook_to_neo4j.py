"""Import the curated Shijiazhuang workbook into Neo4j without deleting data.

The import is idempotent: nodes and relationships are matched by stable workbook IDs.
Credentials are read from NEO4J_* variables, with the local development defaults used
by business-service/application.yml.
"""

from __future__ import annotations

import argparse
import os
from pathlib import Path
from typing import Any

from neo4j import GraphDatabase
from openpyxl import load_workbook


SHEETS = {
    "地点": ("Location", "地点ID", "name", {"标准名称": "name", "类别": "category", "地址": "address", "区县": "district", "片区": "area", "纬度": "latitude", "经度": "longitude", "历史时期": "historicalPeriod", "简介": "intro", "教学标签": "teachingTags", "审核状态": "reviewStatus", "备注": "notes"}),
    "人物": ("Person", "人物ID", "name", {"姓名": "name", "身份": "identity", "出生日期": "birthDate", "逝世日期": "deathDate", "简介": "intro", "审核状态": "reviewStatus"}),
    "事件": ("Event", "事件ID", "name", {"名称": "name", "类别": "category", "开始时间": "startTime", "结束时间": "endTime", "主要地点": "primaryPlace", "简介": "intro", "审核状态": "reviewStatus"}),
    "思政主题": ("IdeologyTheme", "主题ID", "name", {"名称": "name", "类别": "category", "内涵摘要": "summary", "适用学段": "schoolStages", "适用学科": "subjects", "审核状态": "reviewStatus"}),
    "教学资源": ("TeachingResource", "资源ID", "title", {"标题": "title", "类型": "resourceType", "适用学段": "schoolStages", "学科": "subjects", "建议课时": "suggestedLessons", "教学目标": "objectives", "活动流程": "activityFlow", "核心问题": "coreQuestion", "状态": "status"}),
}

TYPE_LABEL = {"地点": "Location", "人物": "Person", "事件": "Event", "思政主题": "IdeologyTheme", "教学资源": "TeachingResource"}
REL_TYPES = {"发生于/关联地点": "OCCURRED_AT", "参与/关联": "PARTICIPATED_IN", "体现": "EMBODIES", "引用": "REFERENCES", "使用": "USES"}


def rows(ws) -> list[dict[str, Any]]:
    values = list(ws.iter_rows(values_only=True))
    headers = values[0]
    return [dict(zip(headers, row)) for row in values[1:] if any(v is not None for v in row)]


def clean(props: dict[str, Any]) -> dict[str, Any]:
    return {k: v for k, v in props.items() if v is not None and v != ""}


def load_data(path: Path) -> dict[str, Any]:
    wb = load_workbook(path, read_only=True, data_only=True)
    control = {r["实体ID"]: r for r in rows(wb["导入控制"])}
    nodes: list[dict[str, Any]] = []
    valid_ids: set[str] = set()
    for sheet, (label, id_col, _, mapping) in SHEETS.items():
        for row in rows(wb[sheet]):
            node_id = str(row[id_col]).strip()
            if label == "Location":
                rule = control.get(node_id)
                if not rule or str(rule["导入Location"]).upper() != "YES" or rule["图谱状态"] == "EXCLUDED":
                    continue
            props = clean({target: row.get(source) for source, target in mapping.items()})
            props.update({"id": node_id, "sourceDataset": "石家庄红色文化原型资源首批数据"})
            if label == "Location":
                props["publishStatus"] = control[node_id]["图谱状态"]
                props["published"] = control[node_id]["图谱状态"] == "PUBLISHED"
            nodes.append({"label": label, "props": props})
            valid_ids.add(node_id)

    sources = []
    for row in rows(wb["来源_标准化"]):
        sid = str(row["来源ID"]).strip()
        props = clean({"id": sid, "title": row["标题"], "sourceType": row["来源类型"], "publisher": row["发布机构"], "url": row["URL"], "trustLevel": row["可信等级"], "purpose": row["用途"], "accessDate": row["访问日期"], "status": row["状态"], "supportedFields": row["支持字段"], "granularity": row["来源粒度"]})
        sources.append(props)

    relations = []
    for row in rows(wb["知识图谱关系"]):
        start, end = str(row["起点ID"]).strip(), str(row["终点ID"]).strip()
        if start not in valid_ids or end not in valid_ids:
            continue
        relations.append({"id": str(row["关系ID"]).strip(), "start": start, "end": end, "startLabel": TYPE_LABEL[row["起点类型"]], "endLabel": TYPE_LABEL[row["终点类型"]], "relType": REL_TYPES[row["关系类型"]], "originalType": row["关系类型"], "reviewStatus": row["审核状态"], "notes": row["备注"]})

    entity_sources = [{"entity": str(r["实体ID"]).strip(), "source": str(r["来源ID"]).strip(), "role": r["来源角色"], "verification": r["核验说明"]} for r in rows(wb["实体来源关系"]) if str(r["实体ID"]).strip() in valid_ids]
    relation_sources = [{"relation": str(r["关系ID"]).strip(), "source": str(r["来源ID"]).strip(), "notes": r["说明"]} for r in rows(wb["关系来源关系"]) if str(r["关系ID"]).strip() in {x["id"] for x in relations}]
    return {"nodes": nodes, "sources": sources, "relations": relations, "entitySources": entity_sources, "relationSources": relation_sources}


def import_data(driver, data: dict[str, Any]) -> None:
    constraints = [
        "CREATE CONSTRAINT location_workbook_id IF NOT EXISTS FOR (n:Location) REQUIRE n.id IS UNIQUE",
        "CREATE CONSTRAINT person_workbook_id IF NOT EXISTS FOR (n:Person) REQUIRE n.id IS UNIQUE",
        "CREATE CONSTRAINT ideology_theme_workbook_id IF NOT EXISTS FOR (n:IdeologyTheme) REQUIRE n.id IS UNIQUE",
        "CREATE CONSTRAINT teaching_resource_workbook_id IF NOT EXISTS FOR (n:TeachingResource) REQUIRE n.id IS UNIQUE",
        "CREATE CONSTRAINT source_workbook_id IF NOT EXISTS FOR (n:Source) REQUIRE n.id IS UNIQUE",
        "CREATE CONSTRAINT graph_relation_workbook_id IF NOT EXISTS FOR (n:GraphRelation) REQUIRE n.id IS UNIQUE",
    ]
    with driver.session() as session:
        for cypher in constraints:
            session.run(cypher).consume()
        for node in data["nodes"]:
            # Locations also carry :Site for compatibility with existing app queries;
            # people also carry :Hero for the same reason.
            extra = ":Site" if node["label"] == "Location" else (":Hero" if node["label"] == "Person" else "")
            set_clause = f"SET n{extra}, n += $props" if extra else "SET n += $props"
            session.run(f"MERGE (n:{node['label']} {{id: $id}}) {set_clause}", id=node["props"]["id"], props=node["props"]).consume()
        for props in data["sources"]:
            session.run("MERGE (n:Source {id: $id}) SET n += $props", id=props["id"], props=props).consume()
        for rel in data["relations"]:
            query = f"MATCH (a:{rel['startLabel']} {{id:$start}}), (b:{rel['endLabel']} {{id:$end}}) MERGE (a)-[r:{rel['relType']} {{id:$id}}]->(b) SET r.originalType=$originalType, r.reviewStatus=$reviewStatus, r.notes=$notes MERGE (g:GraphRelation {{id:$id}}) SET g.type=$relType, g.originalType=$originalType, g.reviewStatus=$reviewStatus MERGE (g)-[:FROM]->(a) MERGE (g)-[:TO]->(b)"
            session.run(query, **rel).consume()
        session.run("UNWIND $rows AS row MATCH (e {id:row.entity}), (s:Source {id:row.source}) MERGE (e)-[r:SUPPORTED_BY]->(s) SET r.role=row.role, r.verification=row.verification", rows=data["entitySources"]).consume()
        session.run("UNWIND $rows AS row MATCH (g:GraphRelation {id:row.relation}), (s:Source {id:row.source}) MERGE (g)-[r:SUPPORTED_BY]->(s) SET r.notes=row.notes", rows=data["relationSources"]).consume()


def counts(driver) -> dict[str, int]:
    with driver.session() as session:
        result = {}
        for label in ("Location", "Person", "Event", "IdeologyTheme", "TeachingResource", "Source", "GraphRelation"):
            result[label] = session.run(f"MATCH (n:{label}) RETURN count(n) AS c").single()["c"]
        result["publishedLocations"] = session.run("MATCH (n:Location {published:true}) RETURN count(n) AS c").single()["c"]
        result["stagingLocations"] = session.run("MATCH (n:Location {publishStatus:'STAGING'}) RETURN count(n) AS c").single()["c"]
        result["excludedLocationNodes"] = session.run("MATCH (n:Location {id:'L016'}) RETURN count(n) AS c").single()["c"]
        result["semanticRelationships"] = session.run("MATCH ()-[r]->() WHERE r.id STARTS WITH 'REL' RETURN count(r) AS c").single()["c"]
        result["sourceRelationships"] = session.run("MATCH ()-[r:SUPPORTED_BY]->(:Source) RETURN count(r) AS c").single()["c"]
    return result


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()
    data = load_data(args.workbook)
    print({k: len(v) for k, v in data.items()})
    if args.dry_run:
        return
    uri = os.getenv("NEO4J_URI", "bolt://127.0.0.1:7687")
    user = os.getenv("NEO4J_USERNAME", "neo4j")
    password = os.getenv("NEO4J_PASSWORD", "12345678")
    with GraphDatabase.driver(uri, auth=(user, password)) as driver:
        driver.verify_connectivity()
        import_data(driver, data)
        print(counts(driver))


if __name__ == "__main__":
    main()
